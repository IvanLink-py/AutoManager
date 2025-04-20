import datetime

from openpyxl import Workbook
from os import path
import os
from openpyxl.styles import Alignment
from DB import DB
from openpyxl.styles import PatternFill, Font


def report_directory():
    if path.exists('report'):
        return
    os.mkdir("report")


def format_report():
    wb = Workbook()
    ws = wb.active

    ws.append(
        ["Дата", "Маршрут", "Время отправления", "Время отправления по расписанию", "Водитель", "Автобус",
         "Успешно завершен"])

    bold_font = Font(bold=True)

    for rows in ws.iter_rows(min_row=ws.max_row, max_row=ws.max_row, min_col=1):
        for cell in rows:
            cell.font = bold_font

    for row in DB.get_report_data():
        ws.append((row[1].strftime("%d.%m.%Y"), int(row[0]), row[1].strftime("%H:%M"), row[2].strftime("%H:%M"),
                   row[3], row[4], "+" if row[5] else "",))

        if not row[5]:
            for rows in ws.iter_rows(min_row=ws.max_row, max_row=ws.max_row, min_col=1):
                for cell in rows:
                    cell.fill = PatternFill(start_color='00FF9999', end_color='00FF9999', fill_type="solid")

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center')

    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value)))) + 1
    for col, value in dims.items():
        ws.column_dimensions[col].width = value

    report_directory()
    ws.title = "Отчет"
    wb.save(r"report\Отчет о рейсах.xlsx")


def format_report_2():
    driver_data, bus_data = DB.get_report_data_2()

    wb = Workbook()
    ws = wb.active

    ws.append(["Водитель", "Всего рейсов", "Успешных рейсов", "Сорвано рейсов", "Процент успешных рейсов"])

    bold_font = Font(bold=True)

    for rows in ws.iter_rows(min_row=ws.max_row, max_row=ws.max_row, min_col=1):
        for cell in rows:
            cell.font = bold_font

    for driver in driver_data:
        ws.append([driver, driver_data[driver][0], driver_data[driver][1], driver_data[driver][2],
                   f'{round((driver_data[driver][1] / driver_data[driver][0]) * 100, 2)}%'])

    ws.append([" "])
    ws.append(["Автобус", "Всего рейсов", "Успешных рейсов", "Сорвано рейсов", "Процент успешных рейсов"])

    for rows in ws.iter_rows(min_row=ws.max_row, max_row=ws.max_row, min_col=1):
        for cell in rows:
            cell.font = bold_font

    for bus in bus_data:
        ws.append([bus, bus_data[bus][0], bus_data[bus][1], bus_data[bus][2],
                   f'{round((bus_data[bus][1] / bus_data[bus][0]) * 100, 2)}%'])

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center')

    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value)))) + 1
    for col, value in dims.items():
        ws.column_dimensions[col].width = value

    report_directory()
    ws.title = "Отчет"
    wb.save(r"report\Отчет о водителях и автобусах.xlsx")

if __name__ == '__main__':
    DB()
    format_report()
    format_report_2()
